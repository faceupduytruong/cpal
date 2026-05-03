import os, csv, subprocess, tempfile, shutil, difflib, subprocess
from fastapi import FastAPI, Query, Body
from fastapi.middleware.cors import CORSMiddleware
from collections import defaultdict
from datetime import datetime
from typing import Optional
from openpyxl import load_workbook

# --- Load config.env ---
script_dir = os.path.dirname(os.path.abspath(__file__))
config_path = os.path.join(script_dir, "config.env")

def load_env(path):
    if os.path.exists(path):
        with open(path, encoding="utf-8") as f:
            for line in f:
                if "=" in line:
                    key, value = line.strip().split("=", 1)
                    os.environ[key] = value

load_env(config_path)

MEGA_USER1 = os.getenv("MEGA_USER1")
MEGA_PASS1 = os.getenv("MEGA_PASS1")
MEGA_USER2 = os.getenv("MEGA_USER2")
MEGA_PASS2 = os.getenv("MEGA_PASS2")

if not MEGA_USER1 or not MEGA_PASS1 or not MEGA_USER2 or not MEGA_PASS2:
    raise ValueError("Thiếu MEGA_USER1/MEGA_PASS1 hoặc MEGA_USER2/MEGA_PASS2 trong config.env")

# Đường dẫn tới megatools.exe
megatools_exe = r"C:\ProgramData\chocolatey\bin\megatools.exe"

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def format_size(size_str: str) -> str:
    """Chuyển bytes sang MB, làm tròn 2 chữ số thập phân"""
    try:
        size = int(size_str)
        return f"{size/1024/1024:.2f} MB"
    except ValueError:
        return size_str

def run_megatools(user, password):
    cmd = [
        megatools_exe, "ls",
        "--username", user,
        "--password", password,
        "--long"
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8", errors="ignore")
    if result.returncode != 0:
        return []
    return result.stdout.splitlines() if result.stdout else []

@app.get("/feed")
def get_feed(q: str = Query(..., description="Từ khóa tìm kiếm")):
    output1 = run_megatools(MEGA_USER1, MEGA_PASS1)
    output2 = run_megatools(MEGA_USER2, MEGA_PASS2)

    results = []

    for account_name, output in [("Account1", output1), ("Account2", output2)]:
        for line in output:
            if q.lower() in line.lower():
                parts = line.strip().split(maxsplit=6)
                if len(parts) == 7:
                    file_id = parts[0]
                    size = format_size(parts[3])
                    date = parts[4] + " " + parts[5]
                    path = parts[6]
                    results.append({
                        "name": os.path.basename(path) if not path.endswith("/") else path,
                        "path": path,
                        "id": file_id,
                        "size": size,
                        "date": date,
                        "type": "folder" if path.endswith("/") else "file",
                        "account": account_name   # thêm thông tin tài khoản
                    })

    return {"feed": results}

@app.get("/stats")
def get_stats(q: Optional[str] = Query(None, description="Từ khóa tìm kiếm (có thể bỏ trống)")):
    output1 = run_megatools(MEGA_USER1, MEGA_PASS1)
    output2 = run_megatools(MEGA_USER2, MEGA_PASS2)

    folder_sizes = defaultdict(int)
    files_per_year = defaultdict(int)

    for output in [output1, output2]:
        for line in output:
            # nếu có query thì lọc, nếu không thì lấy tất cả
            if q is None or q.lower() in line.lower():
                parts = line.strip().split(maxsplit=6)
                if len(parts) == 7:
                    size_str = parts[3]
                    path = parts[6]
                    try:
                        size = int(size_str)
                    except ValueError:
                        size = 0

                    folder = path.strip("/").split("/")[0] if "/" in path else path
                    folder_sizes[folder] += size

                    date_str = parts[4]
                    try:
                        year = datetime.strptime(date_str, "%Y-%m-%d").year
                        if not path.endswith("/"):
                            files_per_year[year] += 1
                    except Exception:
                        pass

    folder_sizes_mb = {folder: round(sz/1024/1024, 2) for folder, sz in folder_sizes.items()}

    return {
        "folder_sizes": folder_sizes_mb,
        "files_per_year": files_per_year
    }

def export_link(path, user, password):
    cmd = [megatools_exe, "export", "--username", user, "--password", password, path]
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode == 0 and result.stdout:
        return result.stdout.strip().splitlines()[0]
    return None

def download_file_from_mega(path, user, password):
    link = export_link(path, user, password)
    if not link:
        raise ValueError("Không tạo được link chia sẻ cho file")

    tmpdir = tempfile.mkdtemp()
    cmd = [megatools_exe, "dl", link, "--path", tmpdir]
    result = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8", errors="ignore")

    if result.returncode != 0:
        raise RuntimeError(f"Lỗi tải file từ Mega: {result.stderr}")

    # tìm file vừa tải về trong tmpdir
    for root, dirs, files in os.walk(tmpdir):
        for f in files:
            fullpath = os.path.join(root, f)
            print(f"[download_file_from_mega] File đã tải về: {fullpath}")
            return fullpath
    return None

@app.get("/compare")
def compare_files(path1: str = Query(...), path2: str = Query(...)):
    # chạy megatools để lấy nội dung file (ví dụ dùng cat hoặc export)
    # ở đây demo bằng cách đọc file text từ local
    try:
        with open(path1, encoding="utf-8") as f1, open(path2, encoding="utf-8") as f2:
            text1 = f1.readlines()
            text2 = f2.readlines()

        diff = difflib.HtmlDiff().make_table(text1, text2,
                                             fromdesc=path1,
                                             todesc=path2,
                                             context=True,
                                             numlines=3)
        return {"diff_html": diff}
    except Exception as e:
        return {"error": str(e)}

@app.post("/compare_excel_multi")
def compare_excel_multi(paths: list[str] = Body(..., embed=True)):
    files = []
    for p in paths:
        f = download_file_from_mega(p, MEGA_USER1, MEGA_PASS1)
        files.append(f)

    workbooks = [load_workbook(f, data_only=True) for f in files]
    diffs = []

    common_sheets = set(workbooks[0].sheetnames)
    for wb in workbooks[1:]:
        common_sheets &= set(wb.sheetnames)

    for sheetname in common_sheets:
        sheets = [wb[sheetname] for wb in workbooks]
        max_row = max(s.max_row for s in sheets)
        max_col = max(s.max_column for s in sheets)
        for row in range(1, max_row+1):
            for col in range(1, max_col+1):
                values = [s.cell(row=row, column=col).value for s in sheets]
                if not all(v == values[0] for v in values):
                    diffs.append({
                        "sheet": sheetname,
                        "cell": sheets[0].cell(row=row, column=col).coordinate,
                        "values": values
                    })

    # --- Xuất ra Desktop với UTF-8 BOM ---
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    desktop_path = r"C:\Users\Admin\OneDrive\Desktop"
    export_path = os.path.join(desktop_path, f"compare_result_{timestamp}.csv")

    with open(export_path, "w", newline="", encoding="utf-8-sig") as csvfile:
        writer = csv.writer(csvfile)
        header = ["Sheet", "Cell"] + [f"File{i+1}" for i in range(len(paths))]
        writer.writerow(header)
        for diff in diffs:
            row = [diff["sheet"], diff["cell"]] + diff["values"]
            writer.writerow(row)

    # cleanup file tạm
    for f in files:
        try:
            os.remove(f)
            shutil.rmtree(os.path.dirname(f), ignore_errors=True)
        except Exception as e:
            print(f"Lỗi xóa file {f}: {e}")

    return {"files": paths, "diffs": diffs, "export_file": export_path}
