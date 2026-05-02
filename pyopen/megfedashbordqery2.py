import os, subprocess, tempfile, shutil, difflib, subprocess
from fastapi import FastAPI, Query
from fastapi.middleware.cors import CORSMiddleware
from collections import defaultdict
from datetime import datetime
from openpyxl import load_workbook

# --- Load config.env ---
script_dir = os.path.dirname(os.path.abspath(__file__))
config_path = os.path.join(script_dir, "config.env")

def load_env(path):
    if os.path.exists(path):
        with open(path, encoding="utf-8") as f:
            for line in f:
                if "=" in line:
                    key, value = line.strip().split("=", 1)
                    os.environ[key] = value

load_env(config_path)

MEGA_USER1 = os.getenv("MEGA_USER1")
MEGA_PASS1 = os.getenv("MEGA_PASS1")
MEGA_USER2 = os.getenv("MEGA_USER2")
MEGA_PASS2 = os.getenv("MEGA_PASS2")

if not MEGA_USER1 or not MEGA_PASS1 or not MEGA_USER2 or not MEGA_PASS2:
    raise ValueError("Thiếu MEGA_USER1/MEGA_PASS1 hoặc MEGA_USER2/MEGA_PASS2 trong config.env")

# Đường dẫn tới megatools.exe
megatools_exe = r"C:\ProgramData\chocolatey\bin\megatools.exe"

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def format_size(size_str: str) -> str:
    """Chuyển bytes sang MB, làm tròn 2 chữ số thập phân"""
    try:
        size = int(size_str)
        return f"{size/1024/1024:.2f} MB"
    except ValueError:
        return size_str

def run_megatools(user, password):
    cmd = [
        megatools_exe, "ls",
        "--username", user,
        "--password", password,
        "--long"
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8", errors="ignore")
    if result.returncode != 0:
        return []
    return result.stdout.splitlines() if result.stdout else []

@app.get("/feed")
def get_feed(q: str = Query(None, description="Từ khóa tìm kiếm")):
    output1 = run_megatools(MEGA_USER1, MEGA_PASS1)
    output2 = run_megatools(MEGA_USER2, MEGA_PASS2)

    results = []

    for account_name, output in [("Account1", output1), ("Account2", output2)]:
        for line in output:
            # nếu có query thì lọc, nếu không thì lấy tất cả
            if not q or q.lower() in line.lower():
                parts = line.strip().split(maxsplit=6)
                if len(parts) == 7:
                    file_id = parts[0]
                    size = format_size(parts[3])
                    date = parts[4] + " " + parts[5]
                    path = parts[6]
                    results.append({
                        "name": os.path.basename(path) if not path.endswith("/") else path,
                        "path": path,
                        "id": file_id,
                        "size": size,
                        "date": date,
                        "type": "folder" if path.endswith("/") else "file",
                        "account": account_name
                    })

    return {"feed": results}

@app.get("/stats")
def get_stats():
    output1 = run_megatools(MEGA_USER1, MEGA_PASS1)
    output2 = run_megatools(MEGA_USER2, MEGA_PASS2)

    folder_sizes = defaultdict(int)
    files_per_year = defaultdict(int)

    for output in [output1, output2]:
        for line in output:
            parts = line.strip().split(maxsplit=6)
            if len(parts) == 7:
                size_str = parts[3]
                path = parts[6]
                try:
                    size = int(size_str)
                except ValueError:
                    size = 0

                # thống kê dung lượng theo folder gốc
                folder = path.strip("/").split("/")[0] if "/" in path else path
                folder_sizes[folder] += size

                # thống kê số lượng file theo năm
                date_str = parts[4]  # định dạng YYYY-MM-DD
                try:
                    year = datetime.strptime(date_str, "%Y-%m-%d").year
                    if not path.endswith("/"):  # chỉ tính file
                        files_per_year[year] += 1
                except Exception:
                    pass

    # chuyển sang MB và dict
    folder_sizes_mb = {folder: round(sz/1024/1024, 2) for folder, sz in folder_sizes.items()}

    return {
        "folder_sizes": folder_sizes_mb,
        "files_per_year": files_per_year
    }

def export_link(path, user, password):
    cmd = [megatools_exe, "export", "--username", user, "--password", password, path]
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode == 0 and result.stdout:
        return result.stdout.strip().splitlines()[0]
    return None

def download_file_from_mega(path, user, password):
    link = export_link(path, user, password)
    if not link:
        raise ValueError("Không tạo được link chia sẻ cho file")

    tmpdir = tempfile.mkdtemp()
    cmd = [megatools_exe, "dl", link, "--path", tmpdir]  # dùng --path thay vì --to
    result = subprocess.run(cmd, capture_output=True, text=True)

    if result.returncode != 0:
        raise RuntimeError(f"Lỗi tải file từ Mega: {result.stderr}")

    # tìm file vừa tải về trong tmpdir
    for root, dirs, files in os.walk(tmpdir):
        for f in files:
            fullpath = os.path.join(root, f)
            print(f"[download_file_from_mega] File đã tải về: {fullpath}")
            return fullpath
    return None

@app.get("/compare")
def compare_files(path1: str = Query(...), path2: str = Query(...)):
    # chạy megatools để lấy nội dung file (ví dụ dùng cat hoặc export)
    # ở đây demo bằng cách đọc file text từ local
    try:
        with open(path1, encoding="utf-8") as f1, open(path2, encoding="utf-8") as f2:
            text1 = f1.readlines()
            text2 = f2.readlines()

        diff = difflib.HtmlDiff().make_table(text1, text2,
                                             fromdesc=path1,
                                             todesc=path2,
                                             context=True,
                                             numlines=3)
        return {"diff_html": diff}
    except Exception as e:
        return {"error": str(e)}

@app.get("/compare_excel")
def compare_excel(path1: str = Query(...), path2: str = Query(...)):
    file1 = download_file_from_mega(path1, MEGA_USER1, MEGA_PASS1)
    file2 = download_file_from_mega(path2, MEGA_USER1, MEGA_PASS1)

    print(f"[compare_excel] File1 đã tải về: {file1}")
    print(f"[compare_excel] File2 đã tải về: {file2}")

    wb1 = load_workbook(file1, data_only=True)
    wb2 = load_workbook(file2, data_only=True)

    diffs = []
    for sheetname in wb1.sheetnames:
        if sheetname in wb2.sheetnames:
            s1, s2 = wb1[sheetname], wb2[sheetname]
            for row in range(1, max(s1.max_row, s2.max_row) + 1):
                for col in range(1, max(s1.max_column, s2.max_column) + 1):
                    v1 = s1.cell(row=row, column=col).value
                    v2 = s2.cell(row=row, column=col).value
                    if v1 != v2:
                        diffs.append({
                            "sheet": sheetname,
                            "cell": s1.cell(row=row, column=col).coordinate,
                            "file1": v1,
                            "file2": v2
                        })

    print(f"[compare_excel] Tổng số khác biệt: {len(diffs)}")

    # --- Xóa file tạm sau khi so sánh ---
    try:
        os.remove(file1)
        os.remove(file2)
        # Nếu muốn xóa cả thư mục tạm:
        shutil.rmtree(os.path.dirname(file1), ignore_errors=True)
        shutil.rmtree(os.path.dirname(file2), ignore_errors=True)
        print("[compare_excel] Đã xóa file tạm.")
    except Exception as e:
        print(f"[compare_excel] Lỗi khi xóa file tạm: {e}")

    return {"diffs": diffs}
